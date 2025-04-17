from fastapi import FastAPI, Request, HTTPException, Depends
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Optional, Any

import json
import os
import calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import uvicorn

# Create FastAPI app
app = FastAPI(title="Shift Scheduler")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure templates and static files
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# Configuration
DATA_DIR = 'data'
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

ENGINEERS_FILE = os.path.join(DATA_DIR, 'engineers.json')
SCHEDULES_FILE = os.path.join(DATA_DIR, 'schedules.json')
WORKPLACES = ["Studio Hispan", "Studio Press", "Nodal", "Engineer Room"]
SHIFTS = ["Shift 1", "Shift 2", "Shift 3"]

# Initialize data files if they don't exist
if not os.path.exists(ENGINEERS_FILE):
    with open(ENGINEERS_FILE, 'w') as f:
        json.dump([], f)

if not os.path.exists(SCHEDULES_FILE):
    with open(SCHEDULES_FILE, 'w') as f:
        json.dump({}, f)

# Data models
class Engineer(BaseModel):
    name: str
    workplaces: List[str]
    limitations: Dict[str, List[str]] = {}

class EngineerDelete(BaseModel):
    name: str

class ShiftData(BaseModel):
    shift1: Optional[str] = None
    shift2: Optional[str] = None
    shift3: Optional[str] = None

class WorkplaceData(BaseModel):
    data: Dict[str, Dict[str, str]] = {}

class ScheduleRequest(BaseModel):
    year: int
    month: int
    workplaces: Dict[str, Dict[str, Dict[str, str]]] = {}

class ExcelRequest(BaseModel):
    year: int
    month: int

# Helper functions
def load_engineers():
    try:
        with open(ENGINEERS_FILE, 'r') as f:
            return json.load(f)
    except:
        return []

def save_engineers(engineers):
    with open(ENGINEERS_FILE, 'w') as f:
        json.dump(engineers, f)

def load_schedules():
    try:
        with open(SCHEDULES_FILE, 'r') as f:
            return json.load(f)
    except:
        return {}

def save_schedules(schedules):
    with open(SCHEDULES_FILE, 'w') as f:
        json.dump(schedules, f)

# Routes
@app.get("/")
async def read_root(request: Request):
    return templates.TemplateResponse("index.html", {
        "request": request,
        "workplaces": WORKPLACES,
        "shifts": SHIFTS
    })

@app.get("/excel")
async def excel_version(request: Request):
    """
    Alternative version of the main page with improved Excel generation capabilities.
    Use this URL if the main page is having issues.
    """
    return templates.TemplateResponse("index.html", {
        "request": request,
        "workplaces": WORKPLACES,
        "shifts": SHIFTS
    })

@app.get("/excel-generator")
async def excel_generator_page(request: Request):
    """
    Standalone page for generating Excel files from the schedule.
    Use this if the main Excel generation functionality isn't working.
    """
    return templates.TemplateResponse("excel_generator.html", {
        "request": request
    })

@app.get("/api/engineers")
async def get_engineers():
    return load_engineers()

@app.post("/api/engineers")
async def add_engineer(engineer: Engineer):
    engineers = load_engineers()
    
    # Check if updating or adding new
    engineer_exists = False
    for eng in engineers:
        if eng['name'] == engineer.name:
            eng['workplaces'] = engineer.workplaces
            eng['limitations'] = engineer.limitations
            engineer_exists = True
            break
            
    if not engineer_exists:
        engineers.append({
            'name': engineer.name,
            'workplaces': engineer.workplaces,
            'limitations': engineer.limitations
        })
        
    save_engineers(engineers)
    return {"status": "success"}

@app.delete("/api/engineers/{name}")
async def delete_engineer(name: str):
    engineers = load_engineers()
    engineers = [eng for eng in engineers if eng['name'] != name]
    save_engineers(engineers)
    return {"status": "success"}

@app.get("/api/schedule")
async def get_schedule(year: int = None, month: int = None):
    if year is None:
        year = datetime.now().year
    if month is None:
        month = datetime.now().month
    
    schedules = load_schedules()
    period_key = f"{year}-{month}"
    
    if period_key in schedules:
        return schedules[period_key]
    return {}

@app.post("/api/schedule")
async def save_schedule(schedule_data: ScheduleRequest):
    schedules = load_schedules()
    
    # Create key for this month/year if not exists
    period_key = f"{schedule_data.year}-{schedule_data.month}"
    if period_key not in schedules:
        schedules[period_key] = {}
        
    for workplace in WORKPLACES:
        if workplace not in schedules[period_key]:
            schedules[period_key][workplace] = {}
            
        workplace_data = schedule_data.workplaces.get(workplace, {})
        for day in workplace_data:
            schedules[period_key][workplace][day] = workplace_data[day]
    
    save_schedules(schedules)
    return {"status": "success"}

@app.post("/api/generate_excel")
async def generate_excel(excel_request: ExcelRequest):
    year = excel_request.year
    month = excel_request.month
    
    schedules = load_schedules()
    period_key = f"{year}-{month}"
    
    if period_key not in schedules:
        raise HTTPException(status_code=404, detail="No schedule data found for selected period")
    
    # Generate Excel files for each workplace
    excel_files = []
    for workplace in WORKPLACES:
        filename = f"{workplace.replace(' ', '_')}_{year}_{month}.xlsx"
        file_path = os.path.join(DATA_DIR, filename)
        
        # Create Excel with formatting
        create_excel_schedule(file_path, workplace, year, month, schedules[period_key].get(workplace, {}))
        excel_files.append(filename)
    
    return {
        "status": "success",
        "files": excel_files
    }

def create_excel_schedule(file_path, workplace, year, month, schedule_data):
    # Get number of days in the month
    num_days = calendar.monthrange(year, month)[1]
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{workplace} Schedule"
    
    # Define styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    weekend_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, color="FFFFFF")
    centered = Alignment(horizontal='center', vertical='center')
    
    # Create title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"{workplace} - {calendar.month_name[month]} {year}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = centered
    
    # Create headers
    headers = ["Day", "Shift 1", "Shift 2", "Shift 3"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = centered
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Fill in days
    for day in range(1, num_days + 1):
        row = day + 3
        
        # Day column
        day_cell = ws.cell(row=row, column=1)
        day_cell.value = f"{day} - {calendar.day_name[calendar.weekday(year, month, day)]}"
        day_cell.border = border
        day_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Weekend formatting
        is_weekend = calendar.weekday(year, month, day) >= 5  # 5 = Saturday, 6 = Sunday
        if is_weekend:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = weekend_fill
        
        # Fill in shifts
        day_str = str(day)
        if day_str in schedule_data:
            for shift_idx, shift in enumerate(SHIFTS, 1):
                shift_key = f"shift{shift_idx}"
                if shift_key in schedule_data[day_str]:
                    cell = ws.cell(row=row, column=shift_idx + 1)
                    cell.value = schedule_data[day_str][shift_key]
                    cell.border = border
                    cell.alignment = centered
                else:
                    ws.cell(row=row, column=shift_idx + 1).border = border
        else:
            for shift_idx in range(1, 4):
                ws.cell(row=row, column=shift_idx + 1).border = border
    
    # Set row height
    for row in range(1, num_days + 5):
        ws.row_dimensions[row].height = 25
    
    # Save workbook
    wb.save(file_path)

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    file_path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path=file_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)